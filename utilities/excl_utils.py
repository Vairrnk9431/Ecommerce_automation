import openpyxl


def get_row_count(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return   sheet.max_row

def get_col_count(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_column

def read_data(file,sheetname,rows_num,col_num):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return(sheet.cell(rows_num,col_num).value)

def write_data(file,sheetname,rows_num,col_num,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    sheet.cell(rows_num,col_num).value=data
    workbook.save(file)






